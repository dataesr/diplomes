import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from utils import swift
from application.server.main.logger import get_logger

logger = get_logger(__name__)

DATA_PATH = os.getenv("MOUNTED_VOLUME_INSCRITS")


def verif(cor_dict: dict):
    os.chdir(DATA_PATH)
    seuil = 50
    vars1 = ['etablissement_id_paysage', 'Type', 'Année universitaire']
    vars2 = vars1.copy()
    vars2.remove('Année universitaire')

    referentiel = pd.DataFrame(cor_dict['C_ETABLISSEMENTS'])
    referentiel.columns = referentiel.columns.str.lower()

    lib_paysage = referentiel[
        ['id_paysage', 'uo_lib', 'id_paysage_actuel', "date_de_creation", "date_de_fermeture", "anciens_codes_uai"]]
    lib_paysage = lib_paysage.rename(columns={"id_paysage": "etablissement_id_paysage"})
    lib_paysage_actuel = referentiel[['id_paysage', 'uo_lib']]
    lib_paysage_actuel.columns = ["id_paysage_actuel", "uo_lib_actuel"]
    lib_paysage_actuel = lib_paysage_actuel.rename(columns={"uo_lib_actuel": "uo_lib_2022"})
    lib_paysage = pd.merge(lib_paysage, lib_paysage_actuel, how='left', on=['id_paysage_actuel'])

    # Charger le fichier CSV
    file_path = 'verification_diplomes.csv'
    data = pd.read_csv(file_path, sep=';', encoding="utf-8", engine='python')

    # Ajouter une colonne "Type" basée sur la condition
    data['Type'] = data['Nombre de diplômes délivrés'].apply(lambda x: 'Principal' if x > 0 else 'Intermédiaire')

    # Additionner les colonnes "Nombre de diplômes intermédiaires délivrés" et "Nombre de diplômes délivrés"
    # si le type est "Intermédiaire"
    data.loc[data['Type'] == 'Intermédiaire', 'Nombre de diplômes délivrés'] += data[
        'Nombre de diplômes intermédiaires délivrés']

    # Mettre à jour 'DIPLOME_r' avec 'DIPlome_int' où 'Nombre de diplômes intermédiaires délivrés' est supérieur à 0
    data.loc[data['Type'] == 'Intermédiaire', 'DIPLOME_r'] = data['DIPlome_int']

    # Supprimer les colonnes spécifiées
    data.drop(columns=['Nombre de diplômes intermédiaires délivrés', 'DIPlome_int'], inplace=True)

    # Calculer la somme des diplômes délivrés pour chaque combinaison unique
    sum_diplomes = data.groupby(vars1, dropna=False)['Nombre de diplômes délivrés'].sum().reset_index()

    # Transposer "Année universitaire" et ajouter des % de variation
    pivot_diplomes = sum_diplomes.pivot_table(
        index=vars2,
        columns='Année universitaire',
        values='Nombre de diplômes délivrés', dropna=False, aggfunc="sum"
    ).reset_index()

    # Convertir toutes les valeurs des colonnes d'années en numériques
    for year in pivot_diplomes.columns[len(vars2):]:
        pivot_diplomes[year] = pd.to_numeric(pivot_diplomes[year], errors='coerce')

    # Remplacer les valeurs NaN par 0
    pivot_diplomes.fillna(0, inplace=True)

    # Calculer les variations en pourcentage
    years = [col for col in pivot_diplomes.columns if col not in vars1 and not col.endswith('% Change')]

    for i in range(1, len(years)):
        year = years[i]
        prev_year = years[i - 1]
        pivot_diplomes[f'{year} % Change'] = ((pivot_diplomes[year] - pivot_diplomes[prev_year]) / pivot_diplomes[
            prev_year]) * 100

    pivot_diplomes = pivot_diplomes.replace([np.inf, -np.inf], np.nan)

    # Identifier les évolutions importantes
    significant_changes = pivot_diplomes.copy()

    change_columns = [col for col in significant_changes.columns if col.endswith('% Change')]

    for col in change_columns:
        significant_changes[col + '_Significant'] = significant_changes[col].apply(
            lambda x: 'Important' if x < -seuil or x > seuil else 'Non-important')

    change_columns = [col for col in significant_changes.columns if col.endswith('% Change_Significant')]

    # Filtrer pour ne garder que les lignes avec au moins une évolution importante
    important_changes = significant_changes[
        significant_changes.apply(lambda row: any(row[change_columns] == 'Important'), axis=1)]

    important_changes = pd.merge(important_changes, lib_paysage, how='left', on=['etablissement_id_paysage'])
    important_changes["comp_paysage"] = important_changes.apply(
        lambda a: "même établissement" if a["etablissement_id_paysage"] == a[
            "id_paysage_actuel"] else "établissement différent", axis=1)

    # Exporter en Excel avec mise en forme conditionnelle
    wb = Workbook()
    ws = wb.active

    # Écrire les données dans le fichier Excel
    for r in dataframe_to_rows(important_changes, index=False, header=True):
        ws.append(r)

    # Définir les styles de remplissage
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    green_fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')

    # Appliquer la mise en forme conditionnelle
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            col_name = important_changes.columns[cell.column - 1]
            if col_name.endswith('_Significant'):
                significance = cell.value
                year = col_name.split('_Significant')[0].split(' % Change')[0]
                year_col_index = important_changes.columns.get_loc(year) + 1
                if significance == 'Important':
                    ws.cell(row=cell.row, column=year_col_index).fill = red_fill
                    cell.fill = red_fill
                elif significance == 'Non-important':
                    ws.cell(row=cell.row, column=year_col_index).fill = green_fill
                    cell.fill = green_fill

    # Sauvegarder le fichier Excel
    excel_file_path = f"{DATA_PATH}od/evolutions_importantes_sans_diplomr.xlsx"
    wb.save(excel_file_path)
    swift.upload_object_path("sas", excel_file_path)

    logger.debug(f"Les résultats ont été exportés vers {excel_file_path}")
